"""
================================================================================
  BOOTSTRAP INDIRECT TEST — IA → PAYMENT METHOD → TAKEOVER PREMIUM
  Mediation via Multinomial Logit (Stage 1) + OLS (Stage 2)
  Author: generated for MA_Analysis_v8_Fixed_updated model structure
  Date: 2025
================================================================================

GUIDANCE FOR THE READER
─────────────────────────
This code implements a Bootstrap Indirect Test to estimate and
test the statistical significance of the INDIRECT effect of Information Asymmetry
(IA) on the takeover premium VIA the choice of payment method (cash / mixed / stock).

The model structure is a two-stage mediation model (Control Function Approach):
  Stage 1: MNL logit → P(Cash) / P(Mixed) / P(Stock)  [mediator model]
  Stage 2: OLS       → Takeover premium                [outcome model]

The INDIRECT effect is defined as:
  Total_indirect = Σ_g  AME_g × β_g_premium

Where AME_g = share_g × (β_g − β̄) is the "redistribution effect" of IA on
payment category g, and β_g_premium is the premium coefficient for category g.

WHY DOES AME SUM TO ZERO?
  Since Σ_g share_g × (β_g − β̄) = Σ_g share_g × β_g − β̄ × Σ_g share_g
                                      = β̄ − β̄ × 1 = 0
  IA only redistributes probability mass BETWEEN categories — nothing is created, nothing
  disappears. It is a pure redistribution effect.
================================================================================
"""

# ─── 1. IMPORTS ────────────────────────────────────────────────────────────
import warnings
import sys
import numpy as np
import pandas as pd
from sklearn.decomposition import PCA
import statsmodels.api as sm
from statsmodels.discrete.discrete_model import MNLogit
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")  # Suppress convergence warnings in bootstrap

# ─── 2. CONFIGURATION (change here for a different specification) ───────────────────
DATA_FILE        = file_path = "C:\\Data\\final_dataset_imputed_v2-2.xlsx"
OUTPUT_FILE      = file_path = "C:\\Data\\bootstrap_indirect_results.xlsx"

# Variables (matches the model structure in MA_Analysis_v8_Fixed_updated.xlsx)
DEPENDENT_VAR    = "premium"             # Takeover premium (4-week)
TREATMENT_VAR    = ["ia_target",         # Information asymmetry — target
                    "ia_acquiror"]       # Information asymmetry — acquiror
MEDIATOR_VAR     = "payment_method"      # 0=Cash, 1=Mixed, 2=Stock
GROUP_VAR        = "hightech"            # High-Tech classification (0/1)
CONTROL_VARS     = ["hostile",           # Hostile bid
                    "challenged"]        # Competing bidders (>1)

# Bootstrap settings
N_BOOTSTRAP      = 5_000                 # Number of bootstrap iterations (at least 5000)
RANDOM_SEED      = 42                    # Reproducibility
ALPHA            = 0.05                  # Significance level
TOLERANCE        = 2e-4                  # Tolerance for reference verification.
                                         # Given reference values are rounded to 4 decimals.
                                         # Maximum rounding effect = ±0.00005 per factor, ×2 = ±1e-4.
                                         # We set the tolerance to 2e-4 to account for
                                         # cumulative rounding errors in the reference values.

# ─── 3. FIXED REFERENCE VALUES (MUST NOT BE CHANGED OR ESTIMATED) ───────────────
# These values are taken directly from MA_Analysis_v8_Fixed_updated.xlsx
# and are used ONLY for validation and comparison — NEVER in bootstrap estimation.

REFERENCE = {
    # Payment shares (full sample, N=2515)
    "share_cash":  0.5026,
    "share_mixed": 0.2191,
    "share_stock": 0.2783,

    # MNL log-odds coefficients — IA_Target (Table 3A, Mixed vs Cash & Stock vs Cash)
    "beta_mixed_target": 0.3966,
    "beta_stock_target": 0.2885,
    "beta_cash_target":  0.0,    # reference category

    # MNL log-odds coefficients — IA_Acquiror (Table 3A)
    "beta_mixed_acq": -0.6900,
    "beta_stock_acq": -0.9687,
    "beta_cash_acq":   0.0,      # reference category

    # Given AME values (MUST NOT BE RECALCULATED — verification only)
    "AME_target": {"cash": -0.0840, "mixed": 0.0503, "stock": 0.0338},
    "AME_acq":    {"cash":  0.2116, "mixed":-0.0589, "stock":-0.1526},
}

# ─── 4. HELPER FUNCTIONS ──────────────────────────────────────────────────────

def load_data(filepath: str) -> pd.DataFrame:
    """
    Reads raw data from the Excel file and returns a DataFrame.
    Error handling: checks that the file exists and that the correct columns exist.
    """
    print("\n" + "="*72)
    print("  STEP 1: DATA LOADING")
    print("="*72)
    try:
        df = pd.read_excel(filepath)
        print(f"  ✓ File loaded. Number of rows: {len(df)}, Columns: {df.shape[1]}")
    except FileNotFoundError:
        sys.exit(f"  ✗ ERROR: File '{filepath}' not found.")
    except Exception as e:
        sys.exit(f"  ✗ ERROR while loading: {e}")
    return df


def build_analytical_sample(df: pd.DataFrame) -> pd.DataFrame:
    """
    Constructs the analytical sample:
      1. Builds payment dummies and control variables
      2. Computes leverage/cash ratios
      3. Winsorizes at 1%/99%
      4. Removes observations with missing values (listwise deletion)
      5. Computes PCA factors for IA (FIXED — NOT recomputed in bootstrap)
    
    Returns a cleaned DataFrame with N ≈ 2515 observations.
    """
    print("\n" + "="*72)
    print("  STEP 2: CONSTRUCTION OF THE ANALYTICAL SAMPLE")
    print("="*72)

    df = df.copy()

    # 2a) Payment categories (reference: Cash = 0)
    df["stock"] = (df["Consideration Structure"] == "Stock Only").astype(int)
    df["mixed"] = (
        df["Consideration Structure"].isin(["Cash and Stock Combination","Cash and Stock"])
    ).astype(int)
    # If neither stock nor mixed → cash
    df["payment_method"] = 0                      # 0 = Cash (reference category)
    df.loc[df["mixed"] == 1, "payment_method"] = 1 # 1 = Mixed
    df.loc[df["stock"] == 1, "payment_method"] = 2 # 2 = Stock

    # 2b) High-Tech, Hostile, Challenged
    df["hightech"]   = (df["Source"] == "High-Tech").astype(int)
    df["hostile"]    = (df["Deal Attitude"] == "Hostile").astype(int)
    df["challenged"] = (df["Number of Bidders"] > 1).astype(int)

    # 2c) Leverage and cash-flow ratios
    df["acq_leverage"] = (
        df["Acquiror Net Debt Last 12 Months (USD, Millions)"]
        / df["Acquiror Total Assets Last 12 Months (USD, Millions)"]
    )
    df["acq_cash_hold"] = (
        df["Acquiror Cash Last 12 Months (USD, Millions)"]
        / df["Acquiror Total Assets Last 12 Months (USD, Millions)"]
    )
    df["tgt_leverage"] = (
        df["Target Net Debt Last 12 Months (USD, Millions)"]
        / df["Target Total Assets Last 12 Months (USD, Millions)"]
    )
    df["premium"] = df["Premium Paid - 4 Weeks Prior to Announcement"]
    df["year"]    = pd.to_datetime(df["Date Announced"]).dt.year

    # 2d) Listwise deletion on required variables
    required_cols = [
        "acq_leverage", "acq_cash_hold", "tgt_leverage", "premium",
        "z_firm_age_target", "z_log_firm_size_target",
        "z_analyst_target", "z_intangible_ratio_target",
        "z_firm_age_acquiror", "z_log_firm_size_acquiror",
        "z_analyst_acquiror", "z_intangible_ratio_acquiror"
    ]
    missing_cols = [c for c in required_cols if c not in df.columns]
    if missing_cols:
        sys.exit(f"  ✗ ERROR: Missing columns: {missing_cols}")

    n_before = len(df)
    df = df.dropna(subset=required_cols).copy()
    n_after  = len(df)
    print(f"  ✓ Listwise deletion: {n_before} → {n_after} (removed: {n_before-n_after})")

    # 2e) Winsorization at 1%/99% (matches MA_Analysis_v8_Fixed_updated.xlsx Table 1E)
    winsorise_cols = [
        "premium", "acq_leverage", "tgt_leverage", "acq_cash_hold",
        "z_firm_age_target", "z_firm_age_acquiror",
        "z_log_firm_size_target", "z_log_firm_size_acquiror",
        "z_analyst_target", "z_analyst_acquiror",
        "z_intangible_ratio_target", "z_intangible_ratio_acquiror"
    ]
    for col in winsorise_cols:
        lo = df[col].quantile(0.01)
        hi = df[col].quantile(0.99)
        df[col] = df[col].clip(lo, hi)

    print(f"  ✓ Winsorization (1%/99%) applied to {len(winsorise_cols)} variables")
    print(f"  ✓ Analytical sample: N = {len(df)}")

    # 2f) PCA factors for Information Asymmetry
    #     CRITICAL: Factors are computed ONCE on the full analytical
    #     sample and held FIXED across the bootstrap iterations.
    #     Modifying the PCA factors within the bootstrap yields non-comparable results.
    df = _compute_ia_factors(df)

    return df.reset_index(drop=True)


def _compute_ia_factors(df: pd.DataFrame) -> pd.DataFrame:
    """
    Computes PCA-based IA factors with fixed sign conventions:
      −Age  : Younger firm → higher IA
      −Size : Smaller firm → higher IA
      −Anal : Fewer analysts → higher IA
      +Intg : More intangible assets → higher IA

    PCA extracts PC1. The sign of PC1 is corrected so that the HT mean
    exceeds the NHT mean (theory-consistent direction).
    
    NOTE: The loadings are fixed and are NOT recomputed in bootstrap iterations.
    """
    # Sign-corrected input vectors (N×4 matrices)
    target_mat = np.column_stack([
        -df["z_firm_age_target"],
        -df["z_log_firm_size_target"],
        -df["z_analyst_target"],
         df["z_intangible_ratio_target"]
    ])
    acq_mat = np.column_stack([
        -df["z_firm_age_acquiror"],
        -df["z_log_firm_size_acquiror"],
        -df["z_analyst_acquiror"],
         df["z_intangible_ratio_acquiror"]
    ])

    # PCA on the full analytical sample (N=2515)
    pca_t = PCA(n_components=1)
    ia_t_raw = pca_t.fit_transform(target_mat).flatten()

    pca_a = PCA(n_components=1)
    ia_a_raw = pca_a.fit_transform(acq_mat).flatten()

    # Sign check: positive factor = higher IA (HT should have the higher mean)
    ht_mask = df["hightech"].values == 1
    ia_t = ia_t_raw if ia_t_raw[ht_mask].mean() > ia_t_raw[~ht_mask].mean() else -ia_t_raw
    ia_a = ia_a_raw if ia_a_raw[ht_mask].mean() > ia_a_raw[~ht_mask].mean() else -ia_a_raw

    df["ia_target"]  = ia_t
    df["ia_acquiror"] = ia_a

    print(f"\n  PCA — IA_Target  : N={len(ia_t)}, Mean={ia_t.mean():.4f}, Std={ia_t.std():.4f}")
    print(f"  PCA — IA_Acquiror: N={len(ia_a)}, Mean={ia_a.mean():.4f}, Std={ia_a.std():.4f}")
    print(f"  IA_Target  HT={ia_t[ht_mask].mean():.4f}, NHT={ia_t[~ht_mask].mean():.4f}")
    print(f"  IA_Acquiror HT={ia_a[ht_mask].mean():.4f}, NHT={ia_a[~ht_mask].mean():.4f}")

    return df


def validate_inputs(df: pd.DataFrame) -> None:
    """
    Checks that:
      1. All required columns are present
      2. No NaN values remain in key columns
      3. Payment categories have reasonable shares
      4. Empty groups are detected
    """
    print("\n" + "="*72)
    print("  STEP 3: INPUT VALIDATION")
    print("="*72)

    required = ["ia_target", "ia_acquiror", "payment_method", "hightech",
                "premium", "stock", "mixed", "hostile", "challenged",
                "acq_leverage", "acq_cash_hold", "tgt_leverage", "year"]
    missing  = [c for c in required if c not in df.columns]
    if missing:
        sys.exit(f"  ✗ Columns missing after data prep: {missing}")

    # NaN check
    nan_counts = {c: df[c].isna().sum() for c in required}
    nan_found  = {c: v for c, v in nan_counts.items() if v > 0}
    if nan_found:
        print(f"  ⚠ Warning: NaN values found: {nan_found}")
    else:
        print("  ✓ No NaN values in key columns")

    # Payment shares
    shares = df["payment_method"].value_counts(normalize=True).sort_index()
    print(f"  ✓ Payment shares → Cash: {shares.get(0,0):.4f}, "
          f"Mixed: {shares.get(1,0):.4f}, Stock: {shares.get(2,0):.4f}")

    # Groups
    for g in [0, 1, 2]:
        n = (df["payment_method"] == g).sum()
        if n < 10:
            print(f"  ⚠ Warning: Payment category {g} has only {n} observations!")
    
    print(f"  ✓ Validation complete. N = {len(df)}")


# ─── 5. AME CALCULATION FUNCTIONS ─────────────────────────────────────────────

def compute_beta_bar(share_cash: float, share_mixed: float, share_stock: float,
                     beta_cash: float, beta_mixed: float, beta_stock: float) -> float:
    """
    Step 1 in the AME calculation: Weighted average of MNL coefficients.

    β̄ = Σ_g (share_g × β_g)
       = share_cash × β_cash + share_mixed × β_mixed + share_stock × β_stock

    Why weight by shares?
    β̄ represents the average log-odds coefficient in the population.
    The weights (shares) reflect how large a portion of the sample actually
    falls into each payment category.

    β_cash = 0 always (reference category in MNL).
    """
    return share_cash * beta_cash + share_mixed * beta_mixed + share_stock * beta_stock


def compute_ame(share_g: float, beta_g: float, beta_bar: float) -> float:
    """
    Step 2 in the AME calculation: Group-specific redistribution effect.

    AME_g = share_g × (β_g − β̄)

    Interpretation:
      – β_g − β̄ measures how much category g "deviates" from the average.
      – Positive AME: IA increases the probability of category g more than the average.
      – Negative AME: IA decreases the probability of category g relative to the average.
      – The sum Σ_g AME_g = 0 always (zero-sum redistribution).

    Parameters:
      share_g   : Share of observations in category g (fixed, from the full sample)
      beta_g    : MNL coefficient for IA in category g vs. the reference (Cash)
      beta_bar  : Weighted average of all β_g
    """
    return share_g * (beta_g - beta_bar)


def compute_all_ames(shares: dict, betas: dict) -> dict:
    """
    Computes β̄ and all three AME_g for one IA measure.
    
    Parameters:
      shares: {"cash": 0.5026, "mixed": 0.2191, "stock": 0.2783}
      betas:  {"cash": 0.0, "mixed": β_mixed, "stock": β_stock}
    
    Returns a dict with beta_bar, AME_cash, AME_mixed, AME_stock.
    """
    beta_bar = compute_beta_bar(
        shares["cash"], shares["mixed"], shares["stock"],
        betas["cash"],  betas["mixed"],  betas["stock"]
    )
    return {
        "beta_bar":  beta_bar,
        "AME_cash":  compute_ame(shares["cash"],  betas["cash"],  beta_bar),
        "AME_mixed": compute_ame(shares["mixed"], betas["mixed"], beta_bar),
        "AME_stock": compute_ame(shares["stock"], betas["stock"], beta_bar),
    }


def compute_total_indirect(ame_stock: float, ame_mixed: float,
                           beta_stock_prem: float, beta_mixed_prem: float) -> float:
    """
    Computes the TOTAL INDIRECT EFFECT of IA on the premium via the payment choice:

      Total_indirect = AME_stock × β_stock_premium + AME_mixed × β_mixed_premium

    Note that AME_cash does not contribute here because Cash is the reference category
    in the premium regression (β_cash_premium = 0 relative to the intercept).

    The indirect effect measures: how much IA changes the premium INDIRECTLY
    by steering the distribution of payment methods.
    """
    return ame_stock * beta_stock_prem + ame_mixed * beta_mixed_prem


# ─── 6. REFERENCE VERIFICATION ──────────────────────────────────────────────────

def verify_reference_values(ref: dict) -> None:
    """
    Reproduces and verifies the given AME reference values.
    The code manually reconstructs every calculation step and compares it against the
    fixed reference values. Deviations > TOLERANCE are flagged as errors.

    🔒 THESE VALUES ARE NEVER ESTIMATED — THEY ARE USED ONLY FOR VERIFICATION.
    """
    print("\n" + "="*72)
    print("  STEP 4: REFERENCE VALUE VERIFICATION")
    print("="*72)

    shares = {
        "cash":  ref["share_cash"],
        "mixed": ref["share_mixed"],
        "stock": ref["share_stock"]
    }

    # ── IA Target ──────────────────────────────────────────────────────────
    print("\n  Reference calculation — IA Target:")
    print(f"  {'─'*60}")

    beta_bar_t = compute_beta_bar(
        shares["cash"], shares["mixed"], shares["stock"],
        ref["beta_cash_target"], ref["beta_mixed_target"], ref["beta_stock_target"]
    )
    print(f"\n  β̄ = {shares['cash']}×0 + {shares['mixed']}×{ref['beta_mixed_target']}"
          f" + {shares['stock']}×{ref['beta_stock_target']}")
    print(f"     = {shares['cash']*0:.4f} + {shares['mixed']*ref['beta_mixed_target']:.4f}"
          f" + {shares['stock']*ref['beta_stock_target']:.4f}")
    print(f"     = {beta_bar_t:.4f}  (expected 0.1672)")

    ame_stock_t = compute_ame(shares["stock"], ref["beta_stock_target"], beta_bar_t)
    ame_mixed_t = compute_ame(shares["mixed"], ref["beta_mixed_target"], beta_bar_t)
    ame_cash_t  = compute_ame(shares["cash"],  ref["beta_cash_target"],  beta_bar_t)

    print(f"\n  AME_stock = {shares['stock']} × ({ref['beta_stock_target']} − {beta_bar_t:.4f})"
          f" = {ame_stock_t:.4f}  (expected +0.0338)")
    print(f"  AME_mixed = {shares['mixed']} × ({ref['beta_mixed_target']} − {beta_bar_t:.4f})"
          f" = {ame_mixed_t:.4f}  (expected +0.0503)")
    print(f"  AME_cash  = {shares['cash']} × (0 − {beta_bar_t:.4f})"
          f" = {ame_cash_t:.4f}  (expected −0.0840)")

    _check_match("AME_stock_target", ame_stock_t, ref["AME_target"]["stock"])
    _check_match("AME_mixed_target", ame_mixed_t, ref["AME_target"]["mixed"])
    _check_match("AME_cash_target",  ame_cash_t,  ref["AME_target"]["cash"])

    sum_t = ame_stock_t + ame_mixed_t + ame_cash_t
    print(f"\n  Sum AME_target = {sum_t:.6f}  ({'≈ 0 ✓' if abs(sum_t) < 0.001 else '≠ 0 ✗'})")

    # ── IA Acquiror ────────────────────────────────────────────────────────
    print("\n  Reference calculation — IA Acquiror:")
    print(f"  {'─'*60}")

    beta_bar_a = compute_beta_bar(
        shares["cash"], shares["mixed"], shares["stock"],
        ref["beta_cash_acq"], ref["beta_mixed_acq"], ref["beta_stock_acq"]
    )
    print(f"\n  β̄ = {shares['cash']}×0 + {shares['mixed']}×({ref['beta_mixed_acq']})"
          f" + {shares['stock']}×({ref['beta_stock_acq']})")
    print(f"     = {shares['cash']*0:.4f} + {shares['mixed']*ref['beta_mixed_acq']:.4f}"
          f" + {shares['stock']*ref['beta_stock_acq']:.4f}")
    print(f"     = {beta_bar_a:.4f}")

    ame_stock_a = compute_ame(shares["stock"], ref["beta_stock_acq"], beta_bar_a)
    ame_mixed_a = compute_ame(shares["mixed"], ref["beta_mixed_acq"], beta_bar_a)
    ame_cash_a  = compute_ame(shares["cash"],  ref["beta_cash_acq"],  beta_bar_a)

    print(f"\n  AME_stock = {shares['stock']} × ({ref['beta_stock_acq']} − {beta_bar_a:.4f})"
          f" = {ame_stock_a:.4f}  (expected −0.1526)")
    print(f"  AME_mixed = {shares['mixed']} × ({ref['beta_mixed_acq']} − {beta_bar_a:.4f})"
          f" = {ame_mixed_a:.4f}  (expected −0.0589)")
    print(f"  AME_cash  = {shares['cash']} × (0 − {beta_bar_a:.4f})"
          f" = {ame_cash_a:.4f}  (expected +0.2116)")

    _check_match("AME_stock_acq", ame_stock_a, ref["AME_acq"]["stock"])
    _check_match("AME_mixed_acq", ame_mixed_a, ref["AME_acq"]["mixed"])
    _check_match("AME_cash_acq",  ame_cash_a,  ref["AME_acq"]["cash"])

    sum_a = ame_stock_a + ame_mixed_a + ame_cash_a
    print(f"\n  Sum AME_acq = {sum_a:.6f}  ({'≈ 0 ✓' if abs(sum_a) < 0.001 else '≠ 0 ✗'})")


def _check_match(name: str, computed: float, given: float) -> None:
    """
    Prints the verification status for a single value.
    
    NOTE: Given reference values are rounded to 4 decimals in the Excel file.
    Deviations within TOLERANCE are to be expected and are due to
    rounding error in the source — NOT a calculation error in this code.
    """
    diff   = computed - given
    status = "✔ matches" if abs(diff) < TOLERANCE else f"✖ DEVIATES (diff={diff:.2e})"
    note   = "" if abs(diff) < TOLERANCE else " [NOTE: deviation > tolerance — check source values]"
    print(f"    [{status}] {name}: computed={computed:.6f}, given={given:.4f}, "
          f"diff={diff:.2e}{note}")


# ─── 7. BOOTSTRAP ITERATION ──────────────────────────────────────────────────

def bootstrap_iteration(df: pd.DataFrame, shares: dict, year_dummies: list) -> dict | None:
    """
    A single bootstrap iteration. Draws a sample with replacement (N=N),
    estimates Stage 1 (MNL) and Stage 2 (OLS), and computes the AME and the indirect effect.

    WHAT THE BOOTSTRAP DOES:
    The bootstrap simulates the sampling variation that would arise if we had drawn
    a new sample from the population. By repeating this N_BOOTSTRAP
    times, an empirical probability distribution is obtained for each estimate, without
    assumptions about normality. Confidence intervals are computed directly from
    the percentiles of this distribution.

    Parameters:
      df          : Analytical sample
      shares      : Fixed payment shares {"cash":…, "mixed":…, "stock":…}
      year_dummies: List of year-dummy columns

    Returns:
      A dict with all computed values, or None if the MNL did not converge.
    """
    # 7a) Resample with replacement
    boot_idx = np.random.randint(0, len(df), size=len(df))
    df_b     = df.iloc[boot_idx].copy()

    try:
        # ─── STAGE 1: Multinomial Logit (Mediator model) ──────────────────
        # Outcome: payment_method (0=Cash, 1=Mixed, 2=Stock)
        # Predictors: IA + interactions + controls
        df_b["ia_t_ht"] = df_b["ia_target"]  * df_b["hightech"]
        df_b["ia_a_ht"] = df_b["ia_acquiror"] * df_b["hightech"]
        df_b["lev_ht"]  = df_b["acq_leverage"] * df_b["hightech"]
        df_b["csh_ht"]  = df_b["acq_cash_hold"] * df_b["hightech"]

        mnl_vars = ["ia_target", "ia_acquiror", "ia_t_ht", "ia_a_ht",
                    "acq_leverage", "acq_cash_hold", "lev_ht", "csh_ht",
                    "tgt_leverage"]
        X_mnl = sm.add_constant(df_b[mnl_vars].values, has_constant="add")
        y_mnl = df_b["payment_method"].values

        mnl_model  = MNLogit(y_mnl, X_mnl)
        mnl_result = mnl_model.fit(method="bfgs", disp=False, maxiter=500)

        if not mnl_result.mle_retvals.get("converged", True):
            return None

        # Extract IA coefficients from the MNL results
        # params is a numpy array (n_vars, 2):
        #   Row 0 = Intercept/const
        #   Row 1 = ia_target, Row 2 = ia_acquiror, …
        #   Column 0 = Mixed vs Cash (category 1 vs 0)
        #   Column 1 = Stock vs Cash (category 2 vs 0)
        coefs     = mnl_result.params  # numpy array (n_vars, 2)
        beta_mixed_t = float(coefs[1, 0])   # ia_target → Mixed
        beta_stock_t = float(coefs[1, 1])   # ia_target → Stock
        beta_mixed_a = float(coefs[2, 0])   # ia_acquiror → Mixed
        beta_stock_a = float(coefs[2, 1])   # ia_acquiror → Stock

        # Premium coefficients from Stage 2 — Mixed and Stock
        # Index 3 = Stock (β₃), Index 4 = Mixed (β₄) in the premium regression
        # (Constant=0, ia_target=1, ia_acquiror=2, stock=3, mixed=4, …)

        # ─── Compute CF Residuals ─────────────────────────────────────────
        # Predicted probabilities from the Stage 1 MNL on the bootstrap sample
        pred_probs   = mnl_result.predict(X_mnl)          # (N, 3): P(cash), P(mixed), P(stock)
        res_stock_b  = (df_b["stock"].values - pred_probs[:, 2])  # actual − predicted P(Stock)
        res_mixed_b  = (df_b["mixed"].values - pred_probs[:, 1])  # actual − predicted P(Mixed)

        # ─── STAGE 2: OLS Premium Regression (Outcome model) ──────────────
        # Includes CF residuals (λ₁, λ₂) to correct for endogeneity.
        # Includes year fixed effects to control for merger waves.
        df_b["ia_t_ht2"]  = df_b["ia_target"]  * df_b["hightech"]
        df_b["ia_a_ht2"]  = df_b["ia_acquiror"] * df_b["hightech"]
        df_b["stk_ht"]    = df_b["stock"]  * df_b["hightech"]
        df_b["mix_ht"]    = df_b["mixed"]  * df_b["hightech"]
        df_b["res_stock"] = res_stock_b
        df_b["res_mixed"] = res_mixed_b

        ols_base = ["ia_target", "ia_acquiror", "stock", "mixed",
                    "ia_t_ht2", "ia_a_ht2", "stk_ht", "mix_ht",
                    "hostile", "challenged", "res_stock", "res_mixed"]

        # Add year dummies (ref: 2002)
        ols_vars  = ols_base + year_dummies
        X_ols     = sm.add_constant(df_b[ols_vars].values, has_constant="add")
        y_ols     = df_b["premium"].values

        ols_result      = sm.OLS(y_ols, X_ols).fit(cov_type="HC1")
        beta_stock_prem = float(ols_result.params[3])   # Coefficient for Stock
        beta_mixed_prem = float(ols_result.params[4])   # Coefficient for Mixed

        # ─── AME and Indirect Effect ──────────────────────────────────────
        # Step 1: Weighted average (β̄) — FIXED shares, bootstrap β
        betas_t = {"cash": 0.0, "mixed": beta_mixed_t, "stock": beta_stock_t}
        betas_a = {"cash": 0.0, "mixed": beta_mixed_a, "stock": beta_stock_a}

        ames_t  = compute_all_ames(shares, betas_t)
        ames_a  = compute_all_ames(shares, betas_a)

        # Step 2: Total indirect effect (AME × premium markup)
        total_indirect_t = compute_total_indirect(
            ames_t["AME_stock"], ames_t["AME_mixed"], beta_stock_prem, beta_mixed_prem
        )
        total_indirect_a = compute_total_indirect(
            ames_a["AME_stock"], ames_a["AME_mixed"], beta_stock_prem, beta_mixed_prem
        )

        return {
            # IA_Target AME
            "beta_bar_t":        ames_t["beta_bar"],
            "AME_stock_t":       ames_t["AME_stock"],
            "AME_mixed_t":       ames_t["AME_mixed"],
            "AME_cash_t":        ames_t["AME_cash"],
            "total_indirect_t":  total_indirect_t,
            # IA_Acquiror AME
            "beta_bar_a":        ames_a["beta_bar"],
            "AME_stock_a":       ames_a["AME_stock"],
            "AME_mixed_a":       ames_a["AME_mixed"],
            "AME_cash_a":        ames_a["AME_cash"],
            "total_indirect_a":  total_indirect_a,
            # Premium coefficients
            "beta_stock_prem":   beta_stock_prem,
            "beta_mixed_prem":   beta_mixed_prem,
            # MNL coefficients
            "beta_mixed_t_raw":  beta_mixed_t,
            "beta_stock_t_raw":  beta_stock_t,
            "beta_mixed_a_raw":  beta_mixed_a,
            "beta_stock_a_raw":  beta_stock_a,
        }

    except Exception:
        return None  # Convergence failure — iteration skipped


def run_bootstrap(df: pd.DataFrame, shares: dict, n_iter: int = N_BOOTSTRAP) -> pd.DataFrame:
    """
    Runs N_BOOTSTRAP iterations of bootstrap_iteration() with replacement.
    
    WHY BOOTSTRAP FOR THE INDIRECT EFFECT?
    The indirect effect is a PRODUCT of coefficients from two separate models.
    The sampling distribution of a product is generally non-normal, and analytical formulas
    (e.g., the Sobel test) often underestimate the variance. The bootstrap provides a
    distribution-free alternative that works even for complex nonlinear
    combinations of estimates.

    Parameters:
      df      : Analytical sample
      shares  : Fixed payment shares
      n_iter  : Number of bootstrap iterations

    Returns:
      A DataFrame with one row per successful iteration.
    """
    print("\n" + "="*72)
    print(f"  STEP 5: BOOTSTRAP ({n_iter:,} ITERATIONS)")
    print("="*72)

    np.random.seed(RANDOM_SEED)

    # Build year dummies (reference: 2002)
    df = df.copy()
    years_present = sorted(df["year"].unique())
    ref_year      = 2002 if 2002 in years_present else years_present[0]
    for yr in years_present:
        if yr != ref_year:
            df[f"yr_{yr}"] = (df["year"] == yr).astype(int)
    year_dummies = [f"yr_{yr}" for yr in years_present if yr != ref_year]

    results     = []
    failed      = 0
    print_every = max(1, n_iter // 10)

    for i in range(n_iter):
        if (i + 1) % print_every == 0 or i == 0:
            pct = 100 * (i + 1) / n_iter
            print(f"  Iteration {i+1:>6,} / {n_iter:,}  ({pct:5.1f}%)  "
                  f"Successful: {len(results):,}  Failed: {failed:,}")

        result = bootstrap_iteration(df, shares, year_dummies)
        if result is not None:
            results.append(result)
        else:
            failed += 1

    boot_df = pd.DataFrame(results)
    print(f"\n  ✓ Bootstrap complete: {len(boot_df):,} successful, {failed:,} failed")
    return boot_df


# ─── 8. SUMMARY & SIGNIFICANCE TEST ─────────────────────────────────────

def summarize_results(boot_df: pd.DataFrame, ref: dict) -> pd.DataFrame:
    """
    Computes for each estimated quantity:
      – Bootstrap mean
      – Standard error (std of the bootstrap distribution)
      – 95% confidence interval (percentile method)
      – Point estimate (bootstrap mean)
    
    Returns a summary table.
    """
    print("\n" + "="*72)
    print("  STEP 6: SUMMARY OF BOOTSTRAP RESULTS")
    print("="*72)

    cols_of_interest = {
        "total_indirect_t": "Total indirect effect — IA_Target",
        "total_indirect_a": "Total indirect effect — IA_Acquiror",
        "AME_stock_t":      "AME_Stock — IA_Target",
        "AME_mixed_t":      "AME_Mixed — IA_Target",
        "AME_cash_t":       "AME_Cash  — IA_Target",
        "AME_stock_a":      "AME_Stock — IA_Acquiror",
        "AME_mixed_a":      "AME_Mixed — IA_Acquiror",
        "AME_cash_a":       "AME_Cash  — IA_Acquiror",
    }

    rows = []
    for col, label in cols_of_interest.items():
        series = boot_df[col].dropna()
        mean   = series.mean()
        se     = series.std()
        ci_lo  = series.quantile(0.025)
        ci_hi  = series.quantile(0.975)
        rows.append({
            "Estimate":          label,
            "Bootstrap Mean":   round(mean, 6),
            "Std Error":        round(se, 6),
            "95% CI Lower":     round(ci_lo, 6),
            "95% CI Upper":     round(ci_hi, 6),
        })

    summary_df = pd.DataFrame(rows)
    print(f"\n{'─'*72}")
    print(f"  {'Estimate':<45} {'Mean':>10} {'SE':>10} {'CI-lo':>10} {'CI-hi':>10}")
    print(f"{'─'*72}")
    for _, row in summary_df.iterrows():
        print(f"  {row['Estimate']:<45} {row['Bootstrap Mean']:>10.4f} "
              f"{row['Std Error']:>10.4f} {row['95% CI Lower']:>10.4f} "
              f"{row['95% CI Upper']:>10.4f}")
    return summary_df


def test_significance(boot_df: pd.DataFrame) -> pd.DataFrame:
    """
    Computes two-sided p-values for each AME and for the total indirect effect.

    P-VALUE VIA BOOTSTRAP:
      The p-value is computed as the proportion of bootstrap draws that are "more extreme"
      than zero (on the relevant side), multiplied by 2 for a two-sided test:

      p = 2 × min( P(AME ≥ 0), P(AME ≤ 0) )

    Interpretation:
      – If p < 0.05 → the AME is statistically significant at the 5% level
      – A confidence interval that does NOT include zero → significant

    WHAT THE BOOTSTRAP P-VALUE MEASURES:
      It reflects the probability of observing an equally extreme (or more
      extreme) estimate IF the true effect is zero. The bootstrap avoids
      assumptions of normality and is robust for complex nonlinear
      estimates such as indirect effects.
    """
    print("\n" + "="*72)
    print("  STEP 7: SIGNIFICANCE TEST FOR EACH AME")
    print("="*72)

    cols_labels = {
        "total_indirect_t": "Total indirect — IA_Target",
        "total_indirect_a": "Total indirect — IA_Acquiror",
        "AME_stock_t":      "AME_Stock — IA_Target",
        "AME_mixed_t":      "AME_Mixed — IA_Target",
        "AME_cash_t":       "AME_Cash  — IA_Target",
        "AME_stock_a":      "AME_Stock — IA_Acquiror",
        "AME_mixed_a":      "AME_Mixed — IA_Acquiror",
        "AME_cash_a":       "AME_Cash  — IA_Acquiror",
    }

    rows = []
    print(f"\n{'─'*72}")
    for col, label in cols_labels.items():
        series = boot_df[col].dropna()
        n      = len(series)

        # Two-sided p-value via the percentile method
        p_pos  = (series >= 0).mean()   # P(AME ≥ 0)
        p_neg  = (series <= 0).mean()   # P(AME ≤ 0)
        p_val  = 2 * min(p_pos, p_neg)
        p_val  = min(p_val, 1.0)         # cap at 1

        ci_lo  = series.quantile(0.025)
        ci_hi  = series.quantile(0.975)
        sig    = p_val < ALPHA
        sig_stars = "***" if p_val < 0.01 else ("**" if p_val < 0.05 else ("*" if p_val < 0.1 else "n.s."))

        status = f"✔ Significant at the {100*ALPHA:.0f}% level ({sig_stars})" if sig else "✗ Not significant"
        print(f"\n  {label}")
        print(f"    p-value = {p_val:.4f}  {sig_stars}  →  {status}")
        print(f"    95% CI: [{ci_lo:.4f}, {ci_hi:.4f}]  "
              f"({'excl. 0' if ci_lo > 0 or ci_hi < 0 else 'incl. 0'})")

        rows.append({
            "Estimate":    label,
            "N Boot":     n,
            "p-value":    round(p_val, 6),
            "Significance":sig_stars,
            "95% CI Lo":  round(ci_lo, 6),
            "95% CI Hi":  round(ci_hi, 6),
            "Significant (5%)": "Yes" if sig else "No"
        })

    return pd.DataFrame(rows)


def consistency_check(boot_df: pd.DataFrame, ref: dict) -> None:
    """
    Checks that AME_stock + AME_mixed + AME_cash ≈ 0 in every iteration
    and in the reference values. This "zero-sum" property is a mathematical
    identity for redistribution AME and should always hold.
    """
    print("\n" + "="*72)
    print("  STEP 8: CONSISTENCY CHECK (Σ AME ≈ 0)")
    print("="*72)

    # Reference values
    sum_t_ref = sum(ref["AME_target"].values())
    sum_a_ref = sum(ref["AME_acq"].values())
    print(f"\n  Reference — IA_Target:  Σ AME = {sum_t_ref:.6f} "
          f"({'≈ 0 ✓' if abs(sum_t_ref) < 0.001 else '✗'})")
    print(f"  Reference — IA_Acquiror: Σ AME = {sum_a_ref:.6f} "
          f"({'≈ 0 ✓' if abs(sum_a_ref) < 0.001 else '✗'})")

    # Bootstrap average
    for label, s, m, c in [
        ("IA_Target",   "AME_stock_t", "AME_mixed_t", "AME_cash_t"),
        ("IA_Acquiror", "AME_stock_a", "AME_mixed_a", "AME_cash_a"),
    ]:
        row_sums = boot_df[s] + boot_df[m] + boot_df[c]
        mean_sum = row_sums.mean()
        max_abs  = row_sums.abs().max()
        print(f"\n  Bootstrap — {label}: "
              f"Mean(Σ AME) = {mean_sum:.2e}, Max|Σ AME| = {max_abs:.2e}  "
              f"({'≈ 0 ✓' if max_abs < 1e-10 else '✓ numerically correct' if max_abs < 1e-6 else '⚠ check'})")


# ─── 9. EXCEL EXPORT ────────────────────────────────────────────────────────

def export_to_excel(boot_df: pd.DataFrame, summary_df: pd.DataFrame,
                    sig_df: pd.DataFrame, ref: dict, output_path: str) -> None:
    """
    Exports all results to an Excel file with formatting:
      Sheet 1: Summary        — Summary of bootstrap results
      Sheet 2: Significance   — Significance test for each AME
      Sheet 3: Reference      — Verification of reference values
      Sheet 4: Bootstrap Draws — All bootstrap draws (raw data)
    """
    print("\n" + "="*72)
    print(f"  STEP 9: EXPORT TO EXCEL — {output_path}")
    print("="*72)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Summary
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        # Sheet 2: Significance
        sig_df.to_excel(writer, sheet_name="Significance", index=False)

        # Sheet 3: Reference verification
        ref_rows = []
        shares = {"cash": ref["share_cash"], "mixed": ref["share_mixed"], "stock": ref["share_stock"]}

        for ia_label, b_cash, b_mixed, b_stock, ame_ref in [
            ("IA_Target",   ref["beta_cash_target"], ref["beta_mixed_target"],
             ref["beta_stock_target"], ref["AME_target"]),
            ("IA_Acquiror", ref["beta_cash_acq"], ref["beta_mixed_acq"],
             ref["beta_stock_acq"], ref["AME_acq"]),
        ]:
            beta_bar = compute_beta_bar(shares["cash"], shares["mixed"], shares["stock"],
                                        b_cash, b_mixed, b_stock)
            for grp, share, beta, ame_given in [
                ("Cash",  shares["cash"],  b_cash,  ame_ref["cash"]),
                ("Mixed", shares["mixed"], b_mixed, ame_ref["mixed"]),
                ("Stock", shares["stock"], b_stock, ame_ref["stock"]),
            ]:
                computed = compute_ame(share, beta, beta_bar)
                diff     = computed - ame_given
                match    = "✔" if abs(diff) < TOLERANCE else "✖"
                ref_rows.append({
                    "IA Variable": ia_label, "Category": grp,
                    "β̄":           round(beta_bar, 6),
                    "share_g":     share, "β_g": beta,
                    "AME Computed": round(computed, 6),
                    "AME Given":    ame_given,
                    "Difference":    round(diff, 8),
                    "Match":        match
                })

        pd.DataFrame(ref_rows).to_excel(writer, sheet_name="Reference Verification", index=False)

        # Sheet 4: Bootstrap draws (all iterations)
        boot_df.to_excel(writer, sheet_name="Bootstrap Draws", index=True)

    # Formatting via openpyxl
    wb = openpyxl.load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F497D")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    body_font   = Font(name="Arial", size=10)

    for ws in wb.worksheets:
        # Format header row
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal="center")
        # Format data rows and column widths
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    wb.save(output_path)
    print(f"  ✓ Exported to '{output_path}'")


# ─── 10. MAIN FUNCTION ─────────────────────────────────────────────────────

def main():
    """
    Orchestrates the entire analysis in 9 steps:
      1. Data loading
      2. Construction of the analytical sample
      3. Validation
      4. Reference value verification
      5. Bootstrap
      6. Summary
      7. Significance test
      8. Consistency check
      9. Export
    """
    print("\n" + "█"*72)
    print("  BOOTSTRAP INDIRECT TEST — IA → PAYMENT CHOICE → TAKEOVER PREMIUM")
    print("█"*72)

    # ── 1. Loading ──────────────────────────────────────────────────────
    df_raw = load_data(DATA_FILE)

    # ── 2. Data prep and analytical sample ──────────────────────────────
    df = build_analytical_sample(df_raw)

    # ── 3. Validation ─────────────────────────────────────────────────────
    validate_inputs(df)

    # ── 4. Reference value verification ──────────────────────────────────
    verify_reference_values(REFERENCE)

    # Payment shares (fixed, computed once from the analytical sample)
    shares = {
        "cash":  float((df["payment_method"] == 0).mean()),
        "mixed": float((df["payment_method"] == 1).mean()),
        "stock": float((df["payment_method"] == 2).mean()),
    }
    print(f"\n  Confirmed shares: Cash={shares['cash']:.4f}, "
          f"Mixed={shares['mixed']:.4f}, Stock={shares['stock']:.4f}")

    # ── 5. Bootstrap ──────────────────────────────────────────────────────
    boot_df = run_bootstrap(df, shares, N_BOOTSTRAP)

    # Check that we have enough successful iterations
    if len(boot_df) < 100:
        sys.exit(f"  ✗ ERROR: Only {len(boot_df)} successful iterations. Check data/model.")

    # ── 6. Summary ─────────────────────────────────────────────────────
    summary_df = summarize_results(boot_df, REFERENCE)

    # ── 7. Significance test ────────────────────────────────────────────────
    sig_df = test_significance(boot_df)

    # ── 8. Consistency check ─────────────────────────────────────────────
    consistency_check(boot_df, REFERENCE)

    # ── 9. Export ─────────────────────────────────────────────────────────
    export_to_excel(boot_df, summary_df, sig_df, REFERENCE, OUTPUT_FILE)

    # ── Final report ───────────────────────────────────────────────────────
    print("\n" + "█"*72)
    print("  ANALYSIS COMPLETE")
    print("█"*72)
    print(f"\n  Bootstrap iterations: {N_BOOTSTRAP:,} (successful: {len(boot_df):,})")
    print(f"  Random seed: {RANDOM_SEED}")
    print(f"  Significance level: {100*ALPHA:.0f}%")
    print(f"  Result file: {OUTPUT_FILE}")
    print()

    return boot_df, summary_df, sig_df


if __name__ == "__main__":
    boot_df, summary_df, sig_df = main()
